import os
import django

os.environ["DJANGO_SETTINGS_MODULE"] = 'mysite.settings'
django.setup()

from openpyxl import load_workbook
from core.models import Room

workbook = load_workbook(filename="roomsExcel.xlsx")
sheet = workbook['DB Rooms']
for i in range(2, 352):
    if i==180:
        data = {}
    data = {}
    if sheet.cell(i, 2).value is not None:
        data['name'] = str(sheet.cell(i, 2).value)
    else:
        continue

    if sheet.cell(i, 3).value is not None:
        data['description'] = sheet.cell(i, 3).value
    else:
        continue

    if sheet.cell(i, 4).value is not None:
        data['address'] = sheet.cell(i, 4).value
    else:
        continue

    if sheet.cell(i, 5).value is not None:
        data['website'] = sheet.cell(i, 5).value
    else:
        continue

    if sheet.cell(i, 6).value is not None:
        data['telephone_num'] = sheet.cell(i, 6).value
    else:
        continue

    if sheet.cell(i, 8).value is not None:
        data['owner'] = sheet.cell(i, 8).value
    else:
        continue

    if sheet.cell(i, 9).value is not None:
        data['duration'] = sheet.cell(i, 9).value

    if sheet.cell(i, 10).value is not None:
        data['is_kids'] = sheet.cell(i, 10).value

    if sheet.cell(i, 11).value is not None:
        data['is_culinary'] = sheet.cell(i, 11).value

    if sheet.cell(i, 12).value is not None:
        data['is_pregnant'] = sheet.cell(i, 12).value

    if sheet.cell(i, 13).value is not None:
        data['is_deaf'] = sheet.cell(i, 13).value

    if sheet.cell(i, 14).value is not None:
        data['minimal_people_amount'] = sheet.cell(i, 14).value
    else:
        continue

    if sheet.cell(i, 15).value is not None:
        data['maximal_people_amount'] = sheet.cell(i, 15).value
    else:
        continue

    if sheet.cell(i, 16).value is not None:
        data['city'] = sheet.cell(i, 16).value
    else:
        continue

    if sheet.cell(i, 17).value is not None:
        data['room_tile_image'] = sheet.cell(i, 17).value

    if sheet.cell(i, 18).value is not None:
        data['large_image'] = sheet.cell(i, 18).value

    try:
        room = Room.objects.create(**data)
        room.save()
        print(room)
    except:
        print("problem with: " ,data)



